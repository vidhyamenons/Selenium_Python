import openpyxl


class homedata:

    test_homepg_data = [{"firstname":"Rahul", "secondname":"shetty", "gender":"Male"}, {"firstname":"Anshika", "secondname":"shetty", "gender":"Female"}]
    @staticmethod
    def excelData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\Rajeev\\Desktop\\Docs\\vsm\\PYTHON SELENIUM\\Test Case Data.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row+1): #to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column+1): #to get coloumns
                    #Dict["firstname"] = "Rahul"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]

